from openai import OpenAI
from system_prompts import system_prompt
import pandas as pd
import json
import re
from openpyxl.styles import PatternFill

# 常量定义
OPENROUTER_API_KEY = "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" #换成相应的api key
OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1" #换成相应的base url
MODEL_NAME = "google/gemini-2.0-flash-lite-preview-02-05:free" #换成gpt-4o
EXCEL_INPUT_FILE = "conference_data.xlsx"
EXCEL_OUTPUT_FILE = "merged_conference_data.xlsx"
REQUIRED_COLUMNS = ["主题", "关键人物或公司"]


def preprocess_prompt(prompt_text):
    """
    预处理提示文本，替换特定词汇
    - 将"华为"替换为"企业"
    - 将"Huawei"替换为"company"（不区分大小写）
    
    Args:
        prompt_text (str): 原始提示文本
        
    Returns:
        str: 处理后的提示文本
    """
    # 替换中文词汇
    processed_text = prompt_text.replace("华为", "企业")
    
    # 不区分大小写替换英文词汇
    pattern = re.compile(r'Huawei', re.IGNORECASE)
    processed_text = pattern.sub('company', processed_text)
    
    return processed_text


def read_conference_data(file_path=EXCEL_INPUT_FILE):
    """
    从Excel文件中读取"主题"和"关键人物或公司"字段，
    返回主题数组和包含Excel中每一行数据的DataFrame。
    
    Args:
        file_path (str): Excel文件路径
        
    Returns:
        tuple: (主题数组, 包含每行数据的DataFrame)，如果出错则返回(None, None)
    """
    try:
        df = pd.read_excel(file_path)
        
        # 检查必要的列是否存在
        for col in REQUIRED_COLUMNS:
            if col not in df.columns:
                print(f"错误：Excel文件中缺少'{col}'列")
                return None, None
        
        # 提取主题列作为数组
        topics = df['主题'].tolist()
        
        return topics, df
    
    except Exception as e:
        print(f"读取Excel文件时出错：{e}")
        return None, None


def create_openai_client(api_key=OPENROUTER_API_KEY, base_url=OPENROUTER_BASE_URL):
    """
    创建OpenAI客户端
    """
    return OpenAI(api_key=api_key, base_url=base_url)


def generate_model_response(client, system_prompt, user_prompt, temperature=0, model=MODEL_NAME):
    """
    生成模型响应
    
    Args:
        client (OpenAI): OpenAI客户端
        system_prompt (str): 系统提示
        user_prompt (str): 用户提示
        temperature (float): 温度参数
        model (str): 模型名称
        
    Returns:
        Completion: 模型完成对象
    """
    return client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        temperature=temperature
    )


def generate_verification_response(client, system_prompt, user_prompt, assistant_response, 
                                  verification_prompt, temperature=0, model=MODEL_NAME):
    """
    生成验证响应
    
    Args:
        client (OpenAI): OpenAI客户端
        system_prompt (str): 系统提示
        user_prompt (str): 用户提示
        assistant_response (str): 助手响应
        verification_prompt (str): 验证提示
        temperature (float): 温度参数
        model (str): 模型名称
        
    Returns:
        Completion: 模型完成对象
    """
    return client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
            {"role": "assistant", "content": assistant_response},
            {"role": "user", "content": verification_prompt}, 
        ],
        temperature=temperature
    )


def parse_json_from_response(content):
    """
    从模型响应中解析JSON数据
    
    Args:
        content (str): 模型响应内容
        
    Returns:
        dict/list: 解析后的JSON数据，如果解析失败则返回None
    """
    # 尝试直接解析JSON
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        # 如果直接解析失败，尝试提取JSON部分
        json_match = re.search(r'(\[.*\]|\{.*\})', content, re.DOTALL)
        if json_match:
            try:
                return json.loads(json_match.group(0))
            except json.JSONDecodeError:
                # 如果仍然失败，进行更复杂的清理
                cleaned_content = content.replace('\n', ' ').replace('```json', '').replace('```', '')
                try:
                    return json.loads(cleaned_content)
                except json.JSONDecodeError:
                    print("无法解析JSON内容，请检查模型输出格式")
                    return None
        else:
            print("无法在输出中找到JSON内容")
            return None


def process_grouped_data(groups, conference_data, client):
    """
    处理分组结果，从conference_data中提取相应的组，
    为每组生成一个子提示字符串，并将每个子提示单独发送给模型处理
    
    Args:
        groups (list): 分组后的索引列表，每个分组是一个包含索引的列表
        conference_data (DataFrame): 包含会议数据的DataFrame
        client (OpenAI): OpenAI客户端
        
    Returns:
        list: 每个分组的处理结果列表
    """
    if not groups or not isinstance(groups, list):
        print("分组数据格式不正确")
        return []
    
    # 为每个分组生成子提示并处理
    group_results = []
    for i, group in enumerate(groups):
        if not isinstance(group, list):
            print(f"跳过格式不正确的分组: {group}")
            continue
        
        # 从conference_data中提取该组的行
        group_rows = []
        original_data = []  # 存储原始数据，用于对比
        for idx in group:
            if 0 <= idx < len(conference_data):
                row = conference_data.iloc[idx]
                # 包含行的所有字段，每个字段前换行
                row_str = f"{idx}."
                for col in conference_data.columns:
                    row_str += f"\n{col}：{row[col]}"
                group_rows.append(row_str)
                
                # 保存原始数据
                original_data.append({
                    "index": idx,
                    "data": row.to_dict()
                })
            else:
                print(f"警告：索引 {idx} 超出范围，已跳过")
        
        # 生成该组的子提示
        if group_rows:
            group_prompt = f"分组 {i+1}:\n" + "\n\n".join(group_rows)
            
            # 输出处理前的原始数据
            print(f"\n分组 {i+1} 处理前的原始数据:")
            for item in original_data:
                print(f"索引 {item['index']}:")
                for key, value in item['data'].items():
                    print(f"  {key}: {value}")
            
            # 将子提示发送给模型处理
            print(f"\n正在处理分组 {i+1}...")
            group_completion = generate_model_response(client, system_prompt, group_prompt)
            group_result = group_completion.choices[0].message.content
            
            # 保存处理结果
            group_results.append({
                "group_number": i+1,
                "prompt": group_prompt,
                "result": group_result,
                "original_data": original_data  # 添加原始数据
            })
            
            print(f"\n分组 {i+1} 处理结果:\n{group_result}\n")
    
    return group_results


def save_merged_results(group_results, original_data, output_file="merged_conference_data.xlsx"):
    """
    将融合后的分组结果与原始数据中未被融合的行合并，并保存为新的Excel文件
    融合的行将被高亮显示
    
    Args:
        group_results (list): 每个分组的处理结果列表
        original_data (DataFrame): 原始会议数据
        output_file (str): 输出文件路径
        
    Returns:
        bool: 是否成功保存
    """
    try:
        # 创建一个新的DataFrame来存储合并后的数据
        merged_data = []
        
        # 跟踪已处理的原始数据索引
        processed_indices = set()
        
        # 跟踪融合的行索引
        merged_row_indices = []
        current_index = 0
        
        # 处理每个分组的结果
        for group_result_item in group_results:
            # 解析JSON结果
            result_json = parse_json_from_response(group_result_item["result"])
            if result_json:
                # 添加融合后的结果到新数据中
                if isinstance(result_json, list):
                    # 记录每个融合结果的索引
                    for _ in result_json:
                        merged_row_indices.append(current_index)
                        current_index += 1
                    merged_data.extend(result_json)
                else:
                    merged_row_indices.append(current_index)
                    current_index += 1
                    merged_data.append(result_json)
                
                # 记录该分组中的所有原始索引
                prompt = group_result_item["prompt"]
                # 使用更精确的正则表达式来提取索引
                indices = re.findall(r'^(\d+)\.', prompt, re.MULTILINE)
                for idx_str in indices:
                    try:
                        processed_indices.add(int(idx_str))
                    except ValueError:
                        print(f"警告：无法解析索引 {idx_str}")
        
        # 添加未处理的原始数据行
        for idx, row in original_data.iterrows():
            if idx not in processed_indices:
                # 将行数据转换为字典并添加到结果中
                row_dict = row.to_dict()
                merged_data.append(row_dict)
                current_index += 1
        
        # 创建DataFrame并保存为Excel
        merged_df = pd.DataFrame(merged_data)
        
        # 使用openpyxl保存并设置高亮
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            merged_df.to_excel(writer, index=False)
            
            # 获取工作表
            worksheet = writer.sheets['Sheet1']
            
            # 创建黄色填充样式
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # 为融合的行应用高亮
            for row_idx in merged_row_indices:
                # Excel行索引从1开始，并且有一个标题行，所以+2
                excel_row = row_idx + 2
                # 为该行的所有单元格应用高亮
                for col in range(1, len(merged_df.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = yellow_fill
        
        print(f"已成功保存融合后的数据到 {output_file}，融合的行已高亮显示")
        return True
        
    except Exception as e:
        print(f"保存融合后的数据时出错：{e}")
        return False


def group_similar_topics(topics):
    """
    利用字符串匹配，查找主题数组中相同主题的索引。
    只返回有多个相同主题的组，单独的主题不会形成组。
    
    Args:
        topics (list): 主题字符串列表
        
    Returns:
        list: 分组后的索引列表，每个分组是一个包含索引的列表。如果没有可匹配的组，返回None
    """
    # 创建主题到索引的映射
    topic_to_indices = {}
    
    # 遍历所有主题，将相同主题的索引分组
    for i, topic in enumerate(topics):
        if topic in topic_to_indices:
            topic_to_indices[topic].append(i)
        else:
            topic_to_indices[topic] = [i]
    
    # 只保留有多个索引的组（即相同主题出现多次的情况）
    groups = [indices for topic, indices in topic_to_indices.items() if len(indices) > 1]
    
    # 如果没有可匹配的组，返回None
    if not groups:
        return None
    
    return groups


def main():
    """主函数，协调整个流程"""
    # 读取会议数据
    topics, conference_data = read_conference_data()
    
    # 使用字符串匹配查找相同主题的索引
    groups = group_similar_topics(topics)
    if groups is None:
        print("没有找到相同主题的组，无需进行融合处理")
        return
    
    print("基于字符串匹配的分组结果：")
    for i, group in enumerate(groups):
        print(f"分组 {i+1}: {group}")
    
    # 创建OpenAI客户端
    client = create_openai_client()
    
    # 处理分组数据并为每个分组生成结果
    group_results = process_grouped_data(groups, conference_data, client)
    
    # 保存融合后的结果到新的Excel文件
    save_merged_results(group_results, conference_data, EXCEL_OUTPUT_FILE)


if __name__ == "__main__":
    main()